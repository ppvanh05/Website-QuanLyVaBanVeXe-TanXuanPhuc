from __future__ import annotations

import csv
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
TESTCASE_DIR = ROOT / "practices" / "testcases"
OUTPUT_DIR = TESTCASE_DIR / "compiled"
REPORTS_DIR = ROOT / "practices" / "reports"
TEMPLATE_PATH = TESTCASE_DIR / "testcase_template.xlsx"
AUTHOR = "Đỗ Thị Phương"

STANDARD_HEADERS = [
    "TC ID",
    "Module",
    "Risk Level",
    "Test Scenario",
    "Pre-Condition",
    "Test Steps",
    "Test Data",
    "Expected Result",
    "Priority",
    "Status",
    "Actual Result",
    "Design By",
    "Design Date",
    "Execute By",
    "Execute Date",
    "Source File",
]

PALETTE = {
    "navy": "0B2A4A",
    "blue": "0070C0",
    "sky": "00B0F0",
    "orange": "F37021",
    "orange_soft": "FCE4D6",
    "green": "70AD47",
    "green_soft": "E2F0D9",
    "red": "C00000",
    "red_soft": "F4CCCC",
    "yellow": "FFD966",
    "yellow_soft": "FFF2CC",
    "gray": "D9E2F3",
    "gray_dark": "808080",
    "white": "FFFFFF",
    "black": "1F2937",
    "surface": "F7FBFF",
    "border": "B7C9D6",
}

STATUS_ORDER = ["Passed", "Failed", "Skip", "Not Run"]
STATUS_COLORS = {
    "Passed": (PALETTE["green_soft"], "375623"),
    "Failed": (PALETTE["red_soft"], PALETTE["red"]),
    "Skip": (PALETTE["yellow_soft"], "7F6000"),
    "Not Run": (PALETTE["gray"], "1F4E79"),
}
PRIORITY_COLORS = {
    "Critical": ("F4CCCC", "990000"),
    "High": (PALETTE["orange_soft"], "C65911"),
    "Medium": (PALETTE["yellow_soft"], "7F6000"),
    "Normal": ("E2F0D9", "375623"),
    "Low": ("E2F0D9", "375623"),
}
SEVERITY_COLORS = {
    "Critical": ("990000", PALETTE["white"]),
    "High": (PALETTE["red_soft"], PALETTE["red"]),
    "Medium": (PALETTE["yellow_soft"], "7F6000"),
    "Low": ("E2F0D9", "375623"),
}
ISSUE_TYPE_COLORS = {
    "Bug": (PALETTE["red_soft"], PALETTE["red"]),
    "Requirement Gap": (PALETTE["orange_soft"], "C65911"),
    "Test Gap": (PALETTE["yellow_soft"], "7F6000"),
}

BUG_REPORT_HEADERS = [
    "Bug ID",
    "Issue Type",
    "Severity",
    "Priority",
    "Status",
    "Module",
    "Related TC",
    "Title",
    "Steps / Trigger",
    "Actual Result",
    "Expected Result",
    "Evidence",
    "Suspected Area",
    "Recommendation",
    "Regression Scope",
    "Source File",
]

THIN_BORDER = Side(style="thin", color=PALETTE["border"])
BOX_BORDER = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = "".join(ch for ch in name if ch not in r"[]:*?/\\").strip() or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    index = 2
    while cleaned in used:
        suffix = f" {index}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(cleaned)
    return cleaned


def pretty_module_name(path: Path, group: str) -> str:
    stem = path.stem
    stem = stem.replace("testcases_", "")
    stem = stem.replace(f"{group}_", "")
    stem = stem.replace("customer_", "")
    stem = stem.replace("admin_", "")
    return " ".join(part.capitalize() for part in stem.split("_"))


def normalize_status(value: str | None) -> str:
    raw = (value or "").strip()
    key = raw.lower().replace("-", " ").replace("_", " ")
    if key in {"pass", "passed"}:
        return "Passed"
    if key in {"fail", "failed"}:
        return "Failed"
    if key in {"skip", "skipped"}:
        return "Skip"
    if key in {"not run", "notrun", "unexecuted", "unexecute", "todo", ""}:
        return "Not Run"
    return raw


def normalize_priority(value: str | None) -> str:
    raw = (value or "").strip()
    key = raw.lower()
    if key in {"critical", "crit"}:
        return "Critical"
    if key == "high":
        return "High"
    if key in {"medium", "med"}:
        return "Medium"
    if key in {"normal", "medium/normal"}:
        return "Normal"
    if key == "low":
        return "Low"
    return "Medium"


def repair_shifted_source_row(source: Dict[str, str]) -> Dict[str, str]:
    """Repair known CSV rows where an unquoted comma shifts fields left."""
    priority = (source.get("Priority") or "").strip()
    status = normalize_status(source.get("Status"))
    if priority and normalize_priority(priority) == "Medium" and priority.lower() not in {"medium", "med"} and status in STATUS_ORDER:
        repaired = dict(source)
        repaired["Test Scenario"] = f"{source.get('Test Scenario', '').strip()}, {source.get('Pre-Condition', '').strip()}".strip(", ")
        repaired["Pre-Condition"] = (source.get("Test Steps") or "").strip()
        repaired["Test Steps"] = (source.get("Test Data") or "").strip()
        repaired["Test Data"] = (source.get("Expected Result") or "").strip()
        repaired["Expected Result"] = priority
        repaired["Priority"] = source.get("Risk Level") or "Medium"
        return repaired
    return source


def read_csv(path: Path, group: str) -> List[Dict[str, str]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "cp1252"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise UnicodeDecodeError("unknown", b"", 0, 1, f"Cannot decode {path}")

    rows: List[Dict[str, str]] = []
    reader = csv.DictReader(text.splitlines())
    module_name = pretty_module_name(path, group)
    for raw_source in reader:
        source = {
            (key or "").lstrip("\ufeff").strip(): value
            for key, value in raw_source.items()
            if key is not None
        }
        source = repair_shifted_source_row(source)
        row = {header: (source.get(header) or "").strip() for header in STANDARD_HEADERS}
        row["TC ID"] = row["TC ID"] or (source.get("ID") or "").strip()
        row["Module"] = row["Module"] or module_name
        row["Test Scenario"] = row["Test Scenario"] or (source.get("NAME") or "").strip()
        row["Pre-Condition"] = row["Pre-Condition"] or (source.get("PRECONDITION") or "").strip()
        row["Test Steps"] = row["Test Steps"] or (source.get("TEST_STEP") or "").strip()
        row["Test Data"] = row["Test Data"] or (source.get("TEST_DATA") or "").strip()
        row["Expected Result"] = row["Expected Result"] or (source.get("EXPECTED_RESULT") or "").strip()
        row["Priority"] = normalize_priority(row["Priority"] or source.get("PRIORITY"))
        row["Risk Level"] = normalize_priority(row["Risk Level"] or row["Priority"])
        row["Status"] = normalize_status(row["Status"] or source.get("STATUS"))
        row["Design By"] = AUTHOR
        row["Source File"] = path.name
        rows.append(row)
    return rows


def load_group_rows(group: str, files: Iterable[Path]) -> List[Dict[str, str]]:
    data: List[Dict[str, str]] = []
    for path in files:
        data.extend(read_csv(path, group))
    return data


def text_blob(row: Dict[str, str]) -> str:
    fields = [
        "Module",
        "Test Scenario",
        "Pre-Condition",
        "Test Steps",
        "Test Data",
        "Expected Result",
        "Actual Result",
        "Source File",
    ]
    return " ".join(row.get(field, "") for field in fields).lower()


def should_report_issue(row: Dict[str, str]) -> bool:
    status = normalize_status(row.get("Status"))
    source = row.get("Source File", "").lower()
    if status == "Failed":
        return True
    if status == "Skip":
        return True
    if status == "Not Run" and source == "testcases_search_trip.csv":
        return True
    return False


def infer_issue_type(row: Dict[str, str]) -> str:
    status = normalize_status(row.get("Status"))
    blob = text_blob(row)
    if status == "Failed":
        return "Bug"
    if any(term in blob for term in ["known gap", "không có", "khong co", "does not", "not implemented", "không giới hạn"]):
        return "Requirement Gap"
    return "Test Gap"


def infer_severity(row: Dict[str, str]) -> str:
    blob = text_blob(row)
    priority = normalize_priority(row.get("Priority") or row.get("Risk Level"))
    if any(term in blob for term in ["mất tiền", "mất dữ liệu", "payment", "thanh toán sai"]):
        return "Critical"
    if any(term in blob for term in ["xss", "<script", "sanitize", "security"]):
        return "High"
    if any(
        term in blob
        for term in [
            "conflict",
            "trùng",
            "đã bán",
            "daban",
            "giucho",
            "giữ chỗ",
            "passenger",
            "số vé",
            "số ghế",
            "seat",
            "booking",
            "create-order",
            "đặt vé",
            "vehicle",
            "driver",
            "tài xế",
            "nhân viên",
        ]
    ):
        return "High"
    if any(term in blob for term in ["otp", "api", "dropdown", "không dấu", "accent", "responsive"]):
        return "Medium"
    if priority in {"Critical", "High", "Medium", "Low"}:
        return priority
    return "Medium"


def priority_from_severity(severity: str, issue_type: str) -> str:
    if severity == "Critical":
        return "P0"
    if severity == "High":
        return "P1"
    if severity == "Medium":
        return "P2"
    if issue_type == "Bug":
        return "P2"
    return "P3"


def clean_title(row: Dict[str, str]) -> str:
    scenario = row.get("Test Scenario", "").strip()
    if scenario:
        return scenario
    return row.get("TC ID", "Issue").strip() or "Issue"


def suspected_area(row: Dict[str, str]) -> str:
    blob = text_blob(row)
    source = row.get("Source File", "").lower()
    if "blacklist" in source or "từ khóa" in blob or "tu khoa" in blob:
        return (
            "kien-tap/src/app/featured/admin/QuanLyTuKhoaCam/quan-ly-tu-khoa-cam.component.ts; "
            "backend/src/admin/tu-khoa-cam/tu-khoa-cam.service.ts"
        )
    if "dispatch" in source or "lịch trình" in blob or "lich trinh" in blob:
        return (
            "kien-tap/src/app/featured/admin/QuanLyDieuHanh/QuanLyLichTrinh/quan-ly-lich-trinh.component.ts; "
            "backend/src/admin/dieu-hanh/dieu-hanh.service.ts"
        )
    if "forgot_password" in source or "otp" in blob or "quên mật khẩu" in blob:
        return (
            "kien-tap/src/app/featured/customer/auth/forgot-password/forgot-password.component.ts; "
            "kien-tap/src/app/featured/customer/auth/forgot-password/forgot-password.component.html; "
            "backend/src/customer/auth/auth.service.ts"
        )
    if "search_trip" in source or "ghế" in blob or "chuyến" in blob:
        return (
            "kien-tap/src/app/featured/customer/tim-kiem-chuyen-xe/tim-kiem-chuyen-xe.ts; "
            "backend/src/customer/tim-kiem-chuyen-xe/tim-kiem-chuyen-xe.service.ts; "
            "backend/src/customer/thong-tin-don-hang/thong-tin-don-hang.service.ts"
        )
    return "Frontend component and corresponding backend validation/API service for this module."


def recommendation(row: Dict[str, str]) -> str:
    blob = text_blob(row)
    if "xss" in blob or "<script" in blob or "sanitize" in blob:
        return "Kiểm tra và làm sạch từ khóa đã chuẩn hóa; từ chối nội dung rỗng hoặc không an toàn ở cả UI và API; thêm kiểm thử hồi quy cho payload dạng script."
    if "255" in blob or "độ dài" in blob or "length" in blob:
        return "Thêm xác thực độ dài tối đa (max-length) rõ ràng trên cả frontend và backend, trả về lỗi chi tiết thay vì chỉ dựa vào hành vi chặn của UI."
    if ("vehicle" in blob or "xe" in blob) and ("conflict" in blob or "trùng" in blob):
        return "Trước khi tạo hoặc cập nhật lịch trình, hãy truy vấn kiểm tra trùng lịch của cùng một xe và trả về mã lỗi 409 hoặc lỗi xác thực nếu xe đã bận."
    if "driver" in blob or "tài xế" in blob or "nhân viên" in blob:
        return "Áp dụng cùng một cơ chế xác thực trùng lịch cho tài xế/phụ xe được phân công và giữ kiểm tra này ở backend làm nguồn xác thực chuẩn (source of truth)."
    if "completed" in blob or "hoàn thành" in blob or "release" in blob:
        return "Xác định trạng thái rảnh/bận của tài nguyên dựa trên trạng thái hoặc thời gian lịch trình hoặc cập nhật trạng thái tài nguyên khi lịch trình hoàn tất."
    if "otp" in blob and ("disable" in blob or "vô hiệu" in blob):
        return "Ràng buộc trạng thái disabled/loading của nút gửi OTP với biến isSendingOtp và thêm guard return sớm trong hàm sendOtp()."
    if "esc" in blob or "escape" in blob:
        return "Xử lý phím Escape ở cấp độ modal để đóng hoặc reset modal quên mật khẩu một cách nhất quán."
    if "swap" in blob or "hoán đổi" in blob:
        return "Hoán đổi cả hai giá trị đã chọn và dữ liệu hiển thị trên input tìm kiếm, sau đó cập nhật hoặc xóa trạng thái tìm kiếm phụ thuộc một cách nhất quán."
    if "passenger" in blob or "số vé" in blob or "số ghế" in blob:
        return "Xác định quy tắc nghiệp vụ cho tỷ lệ số lượng khách hàng so với số ghế và thực thi kiểm tra này ở cả UI và API validation khi tạo đơn hàng."
    if "giucho" in blob or "giữ chỗ" in blob:
        return "Xử lý các ghế có trạng thái GiuCho đang hoạt động là không khả dụng khi gọi API tạo đơn hàng (create-order), trừ khi ghế giữ chỗ đó thuộc về chính phiên làm việc của khách hàng hiện tại."
    if "dropdown" in blob or "không dấu" in blob or "accent" in blob:
        return "Chuẩn hóa bỏ dấu tiếng Việt cho cả từ khóa tìm kiếm và văn bản của các tùy chọn trước khi thực hiện lọc kết quả tự động hoàn thành (autocomplete/dropdown)."
    if "responsive" in blob or "mobile" in blob:
        return "Bổ sung kiểm thử hồi quy cho kích thước màn hình mobile và khắc phục lỗi tràn chữ hoặc vỡ bố cục đối với sơ đồ ghế và tìm chuyến xe trên mobile."
    if "xác nhận chỗ" in blob or "chọn ghế" in blob:
        return "Đồng bộ hóa thuật ngữ trong testcase và giao diện: sử dụng đúng luồng nút hành động (CTA), và thêm bước chặn điều hướng khi chưa chọn ghế."
    return "Xác nhận lại quy tắc nghiệp vụ mong muốn với Product Owner, sửa lỗi xác thực nguồn, sau đó bổ sung test case Playwright để kiểm thử hồi quy."


def regression_scope(row: Dict[str, str]) -> str:
    source = row.get("Source File", "")
    tc_id = row.get("TC ID", "")
    if "search_trip" in source:
        return f"Chạy file test src/tests/customer/search-trip.spec.ts, đặc biệt là {tc_id}; đồng thời test lại luồng thanh toán đặt vé với các ghế đã bán/đang giữ."
    if "forgot_password" in source:
        return f"Chạy file test src/tests/customer/forgot-password.spec.ts, đặc biệt là {tc_id}; test lại việc gửi lại OTP và đóng modal."
    if "blacklist" in source:
        return f"Chạy file test src/tests/admin/blacklist.spec.ts, đặc biệt là {tc_id}; test lại quyền tạo/sửa/xóa từ khóa cấm."
    if "dispatch" in source:
        return f"Chạy file test src/tests/admin/dispatch.spec.ts, đặc biệt là {tc_id}; test lại việc tạo/sửa lịch trình và phân công tài nguyên."
    return f"Chạy các test case Playwright của file {source}, đặc biệt là {tc_id}."


def evidence(row: Dict[str, str]) -> str:
    status = normalize_status(row.get("Status"))
    return (
        f"CSV: {row.get('Source File', '')}; TC: {row.get('TC ID', '')}; "
        f"Status: {status}; Actual: {row.get('Actual Result', '') or 'Not captured in CSV'}"
    )


def build_bug_reports(group_name: str, rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    reports: List[Dict[str, str]] = []
    sequence = 1
    for row in rows:
        if not should_report_issue(row):
            continue
        issue_type = infer_issue_type(row)
        severity = infer_severity(row)
        report = {
            "Bug ID": f"BUG-{group_name.upper()}-{sequence:03d}",
            "Issue Type": issue_type,
            "Severity": severity,
            "Priority": priority_from_severity(severity, issue_type),
            "Status": normalize_status(row.get("Status")),
            "Module": row.get("Module", ""),
            "Related TC": row.get("TC ID", ""),
            "Title": clean_title(row),
            "Steps / Trigger": row.get("Test Steps", ""),
            "Actual Result": row.get("Actual Result", ""),
            "Expected Result": row.get("Expected Result", ""),
            "Evidence": evidence(row),
            "Suspected Area": suspected_area(row),
            "Recommendation": recommendation(row),
            "Regression Scope": regression_scope(row),
            "Source File": row.get("Source File", ""),
        }
        reports.append(report)
        sequence += 1
    return reports


def apply_title_style(ws, title: str, subtitle: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=PALETTE["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=PALETTE["navy"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=PALETTE["blue"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20


def copy_template_chart_sizes(template_path: Path) -> Dict[str, tuple[float, float]]:
    sizes: Dict[str, tuple[float, float]] = {}
    try:
        wb = load_workbook(template_path)
        ws = wb["Overall"]
        for chart in ws._charts:
            title = getattr(chart, "title", None)
            text = ""
            try:
                text = title.tx.rich.p[0].r[0].t
            except Exception:
                text = type(chart).__name__
            sizes[text] = (chart.width, chart.height)
    except Exception:
        pass
    return sizes


def write_status_legend(ws, start_row: int, start_col: int) -> None:
    ws.cell(start_row, start_col, "LEGEND").font = Font(bold=True, color=PALETTE["white"])
    ws.cell(start_row, start_col).fill = PatternFill("solid", fgColor=PALETTE["blue"])
    ws.cell(start_row, start_col + 1).fill = PatternFill("solid", fgColor=PALETTE["blue"])
    ws.cell(start_row, start_col).border = BOX_BORDER
    ws.cell(start_row, start_col + 1).border = BOX_BORDER

    for idx, status in enumerate(STATUS_ORDER, start=start_row + 1):
        fill, font = STATUS_COLORS[status]
        ws.cell(idx, start_col, status)
        ws.cell(idx, start_col + 1, "Execution status")
        for col in (start_col, start_col + 1):
            cell = ws.cell(idx, col)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(bold=col == start_col, color=font)
            cell.border = BOX_BORDER


def write_kpi_card(ws, cell_range: str, label: str, value, fill: str, font_color: str = PALETTE["white"]) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = f"{label}\n{value}"
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, size=12, color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws[cell_range]:
        for c in row:
            c.border = BOX_BORDER


def scenario_type(value: str | None) -> str:
    text = (value or "").strip()
    if " - " in text:
        return text.split(" - ", 1)[0].strip()
    if "-" in text:
        return text.split("-", 1)[0].strip()
    return text.split(" ", 1)[0].strip() if text else "(blank)"


def attention_count(counter: Counter) -> int:
    return counter.get("Failed", 0) + counter.get("Skip", 0) + counter.get("Not Run", 0)


def pass_rate(counter: Counter) -> float:
    total = sum(counter.get(status, 0) for status in STATUS_ORDER)
    return counter.get("Passed", 0) / total if total else 0


def grouped_status_rows(rows: List[Dict[str, str]], key_name: str) -> List[List[object]]:
    grouped: Dict[str, Counter] = defaultdict(Counter)
    for row in rows:
        key = row.get(key_name, "").strip() or "(blank)"
        grouped[key][row["Status"]] += 1

    output: List[List[object]] = []
    for key, counter in grouped.items():
        total = sum(counter.get(status, 0) for status in STATUS_ORDER)
        output.append(
            [
                key,
                counter.get("Passed", 0),
                counter.get("Failed", 0),
                counter.get("Skip", 0),
                counter.get("Not Run", 0),
                total,
                pass_rate(counter),
                attention_count(counter),
            ]
        )
    output.sort(key=lambda item: (item[7], item[5], item[0]), reverse=True)
    return output


def write_dashboard_section(ws, row_idx: int, title: str, end_col: int = 9) -> int:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=end_col)
    cell = ws.cell(row_idx, 1, title)
    cell.fill = PatternFill("solid", fgColor=PALETTE["blue"])
    cell.font = Font(bold=True, color=PALETTE["white"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BOX_BORDER
    ws.row_dimensions[row_idx].height = 22
    return row_idx + 1


def write_dashboard_table(
    ws,
    row_idx: int,
    headers: List[str],
    values: List[List[object]],
    percent_cols: set[int] | None = None,
    long_rows: bool = False,
) -> int:
    percent_cols = percent_cols or set()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row_idx, col_idx, header)
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(bold=True, color=PALETTE["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX_BORDER

    for offset, row_values in enumerate(values, start=1):
        excel_row = row_idx + offset
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(excel_row, col_idx, value)
            cell.border = BOX_BORDER
            cell.alignment = Alignment(
                horizontal="left" if col_idx in {1, 2, 3, 6, 7, 8} else "center",
                vertical="top",
                wrap_text=True,
            )
            if excel_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFD")
            if col_idx in percent_cols:
                cell.number_format = "0.0%"
            if col_idx == 1 and value in STATUS_COLORS:
                fill, font = STATUS_COLORS[str(value)]
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(bold=True, color=font)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(value, str) and value.startswith("#'"):
                cell.hyperlink = value
                cell.value = "Open"
                cell.style = "Hyperlink"
        ws.row_dimensions[excel_row].height = 58 if long_rows else 24
    return row_idx + len(values) + 1


def add_dashboard_details(ws, group_name: str, rows: List[Dict[str, str]], start_row: int) -> None:
    totals = Counter(row["Status"] for row in rows)
    total = len(rows)
    executed = total - totals.get("Not Run", 0)
    open_items = totals.get("Failed", 0) + totals.get("Skip", 0) + totals.get("Not Run", 0)
    modules = {row.get("Module", "") for row in rows if row.get("Module")}
    sources = {row.get("Source File", "") for row in rows if row.get("Source File")}
    high_priority = [row for row in rows if normalize_priority(row.get("Priority")) in {"Critical", "High"}]
    high_risk = [row for row in rows if normalize_priority(row.get("Risk Level")) in {"Critical", "High"}]

    row_idx = write_dashboard_section(ws, start_row, "DETAILED EXECUTION OVERVIEW")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
    note = ws.cell(row_idx, 1)
    note.value = "Detailed summary generated from all source CSV rows: status, module, priority, risk, scenario type, source file, and open-item list."
    note.font = Font(italic=True, color=PALETTE["gray_dark"])
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note.border = BOX_BORDER
    row_idx += 2

    row_idx = write_dashboard_section(ws, row_idx, "QUALITY SNAPSHOT")
    snapshot = [
        ["Total Test Cases", total, "Executed", executed, "Pass Rate", f"{totals.get('Passed', 0) / total:.1%}" if total else "0.0%"],
        ["Passed", totals.get("Passed", 0), "Failed", totals.get("Failed", 0), "Open Items", open_items],
        ["Skip", totals.get("Skip", 0), "Not Run", totals.get("Not Run", 0), "Source Files", len(sources)],
        ["High/Critical Priority", len(high_priority), "High/Critical Priority Open", sum(1 for item in high_priority if item["Status"] != "Passed"), "Modules", len(modules)],
        ["High/Critical Risk", len(high_risk), "High/Critical Risk Open", sum(1 for item in high_risk if item["Status"] != "Passed"), "Designer", AUTHOR],
        ["Workbook Group", group_name.title(), "Generated From", "practices/testcases", "Dashboard Scope", "All rows"],
    ]
    row_idx = write_dashboard_table(ws, row_idx, ["Metric", "Value", "Metric", "Value", "Metric", "Value"], snapshot) + 1

    module_rows = grouped_status_rows(rows, "Module")
    attention_modules = [item for item in module_rows if item[7] > 0] or module_rows[:10]
    row_idx = write_dashboard_section(ws, row_idx, "MODULES NEEDING ATTENTION")
    row_idx = write_dashboard_table(
        ws,
        row_idx,
        ["Module", "Passed", "Failed", "Skip", "Not Run", "Total", "Pass Rate", "Attention"],
        attention_modules[:15],
        {7},
    ) + 1

    row_idx = write_dashboard_section(ws, row_idx, "PRIORITY SUMMARY")
    row_idx = write_dashboard_table(
        ws,
        row_idx,
        ["Priority", "Passed", "Failed", "Skip", "Not Run", "Total", "Pass Rate", "Attention"],
        grouped_status_rows(rows, "Priority"),
        {7},
    ) + 1

    row_idx = write_dashboard_section(ws, row_idx, "RISK SUMMARY")
    row_idx = write_dashboard_table(
        ws,
        row_idx,
        ["Risk Level", "Passed", "Failed", "Skip", "Not Run", "Total", "Pass Rate", "Attention"],
        grouped_status_rows(rows, "Risk Level"),
        {7},
    ) + 1

    scenario_counters: Dict[str, Counter] = defaultdict(Counter)
    for source_row in rows:
        scenario_counters[scenario_type(source_row.get("Test Scenario"))][source_row["Status"]] += 1
    scenario_rows: List[List[object]] = []
    for label, counter in scenario_counters.items():
        subtotal = sum(counter.get(status, 0) for status in STATUS_ORDER)
        scenario_rows.append(
            [
                label,
                counter.get("Passed", 0),
                counter.get("Failed", 0),
                counter.get("Skip", 0),
                counter.get("Not Run", 0),
                subtotal,
                pass_rate(counter),
                attention_count(counter),
            ]
        )
    scenario_rows.sort(key=lambda item: (item[5], item[0]), reverse=True)
    row_idx = write_dashboard_section(ws, row_idx, "SCENARIO TYPE SUMMARY")
    row_idx = write_dashboard_table(
        ws,
        row_idx,
        ["Scenario Type", "Passed", "Failed", "Skip", "Not Run", "Total", "Pass Rate", "Attention"],
        scenario_rows,
        {7},
    ) + 1

    row_idx = write_dashboard_section(ws, row_idx, "SOURCE FILE COVERAGE")
    row_idx = write_dashboard_table(
        ws,
        row_idx,
        ["Source File", "Passed", "Failed", "Skip", "Not Run", "Total", "Pass Rate", "Attention"],
        grouped_status_rows(rows, "Source File"),
        {7},
    ) + 1

    priority_rank = {"Critical": 0, "High": 1, "Medium": 2, "Normal": 3, "Low": 4}
    status_rank = {"Failed": 0, "Skip": 1, "Not Run": 2}
    attention_rows = [
        (idx, source_row)
        for idx, source_row in enumerate(rows, start=2)
        if source_row.get("Status") in {"Failed", "Skip", "Not Run"}
    ]
    attention_rows.sort(
        key=lambda item: (
            status_rank.get(item[1]["Status"], 9),
            priority_rank.get(normalize_priority(item[1].get("Priority")), 9),
            priority_rank.get(normalize_priority(item[1].get("Risk Level")), 9),
            item[1].get("TC ID", ""),
        )
    )
    attention_values = [
        [
            item["Status"],
            item.get("TC ID", ""),
            item.get("Module", ""),
            item.get("Priority", ""),
            item.get("Risk Level", ""),
            item.get("Test Scenario", ""),
            item.get("Actual Result", ""),
            item.get("Source File", ""),
            f"#'All {group_name.title()}'!A{row_number}",
        ]
        for row_number, item in attention_rows[:30]
    ]
    row_idx = write_dashboard_section(ws, row_idx, "ATTENTION ITEMS (FAILED / SKIP / NOT RUN - FIRST 30)")
    write_dashboard_table(
        ws,
        row_idx,
        ["Status", "TC ID", "Module", "Priority", "Risk", "Scenario", "Actual Result", "Source File", "Open Row"],
        attention_values,
        long_rows=True,
    )


def add_dashboard(wb: Workbook, group_name: str, rows: List[Dict[str, str]], chart_sizes: Dict[str, tuple[float, float]]) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    apply_title_style(
        ws,
        f"TXP BUS - {group_name.upper()} TEST CASE DASHBOARD",
        "Consolidated testcase workbook generated from practices/testcases without modifying source CSV files.",
    )

    totals = Counter(row["Status"] for row in rows)
    total = len(rows)
    passed = totals.get("Passed", 0)
    failed = totals.get("Failed", 0)
    skipped = totals.get("Skip", 0)
    not_run = totals.get("Not Run", 0)
    pass_rate = passed / total if total else 0

    ws["A4"] = "META DATA"
    ws["A4"].fill = PatternFill("solid", fgColor=PALETTE["sky"])
    ws["A4"].font = Font(bold=True, color=PALETTE["white"])
    ws["B4"].fill = PatternFill("solid", fgColor=PALETTE["sky"])
    metadata = [
        ("Project", "TXP Bus E2E Testcases"),
        ("Group", group_name.title()),
        ("Generated At", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Source Folder", "practices/testcases"),
    ]
    for idx, (label, value) in enumerate(metadata, start=5):
        ws.cell(idx, 1, label).font = Font(bold=True, color=PALETTE["blue"])
        ws.cell(idx, 2, value).font = Font(bold=True, color=PALETTE["red"] if label == "Group" else PALETTE["black"])
        ws.cell(idx, 1).border = BOX_BORDER
        ws.cell(idx, 2).border = BOX_BORDER

    write_status_legend(ws, 4, 4)

    write_kpi_card(ws, "A11:B13", "TOTAL", total, PALETTE["navy"])
    write_kpi_card(ws, "C11:D13", "PASSED", passed, PALETTE["green"])
    write_kpi_card(ws, "E11:F13", "FAILED", failed, PALETTE["red"])
    write_kpi_card(ws, "G11:H13", "PASS RATE", f"{pass_rate:.1%}", PALETTE["orange"])
    write_kpi_card(ws, "A15:B17", "SKIP", skipped, "BF9000")
    write_kpi_card(ws, "C15:D17", "NOT RUN", not_run, "5B9BD5")
    write_kpi_card(ws, "E15:H17", "REPORT SCOPE", f"{group_name.title()} test case consolidation", PALETTE["blue"])

    status_start = 21
    ws.cell(status_start, 1, "STATUS SUMMARY").font = Font(bold=True, color=PALETTE["white"])
    ws.cell(status_start, 1).fill = PatternFill("solid", fgColor=PALETTE["blue"])
    ws.cell(status_start, 2).fill = PatternFill("solid", fgColor=PALETTE["blue"])
    ws.cell(status_start + 1, 1, "Status")
    ws.cell(status_start + 1, 2, "Count")
    for col in range(1, 3):
        ws.cell(status_start + 1, col).fill = PatternFill("solid", fgColor=PALETTE["navy"])
        ws.cell(status_start + 1, col).font = Font(bold=True, color=PALETTE["white"])
    for offset, status in enumerate(STATUS_ORDER, start=2):
        fill, font = STATUS_COLORS[status]
        ws.cell(status_start + offset, 1, status)
        ws.cell(status_start + offset, 2, totals.get(status, 0))
        for col in range(1, 3):
            ws.cell(status_start + offset, col).fill = PatternFill("solid", fgColor=fill)
            ws.cell(status_start + offset, col).font = Font(bold=col == 1, color=font)
            ws.cell(status_start + offset, col).border = BOX_BORDER

    module_counts: Dict[str, Counter] = defaultdict(Counter)
    for row in rows:
        module_counts[row["Module"]][row["Status"]] += 1

    module_start = 30
    headers = ["Module", "Passed", "Failed", "Skip", "Not Run", "Total"]
    ws.cell(module_start, 1, "MODULE BREAKDOWN").font = Font(bold=True, color=PALETTE["white"])
    ws.cell(module_start, 1).fill = PatternFill("solid", fgColor=PALETTE["blue"])
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(module_start + 1, idx, header)
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(bold=True, color=PALETTE["white"])
        cell.border = BOX_BORDER

    for row_idx, module in enumerate(sorted(module_counts), start=module_start + 2):
        counter = module_counts[module]
        values = [
            module,
            counter.get("Passed", 0),
            counter.get("Failed", 0),
            counter.get("Skip", 0),
            counter.get("Not Run", 0),
            sum(counter.values()),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = BOX_BORDER
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3F8FC")

    details_start = module_start + len(module_counts) + 5
    add_dashboard_details(ws, group_name, rows, details_start)

    for col, width in {"A": 28, "B": 18, "C": 22, "D": 22, "E": 18, "F": 44, "G": 42, "H": 28, "I": 13}.items():
        ws.column_dimensions[col].width = width
    for row in range(4, max(module_start + len(module_counts) + 4, 44)):
        ws.row_dimensions[row].height = 21

    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=status_start + 2, max_row=status_start + 5)
    data = Reference(ws, min_col=2, min_row=status_start + 1, max_row=status_start + 5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Overall Status"
    pie.width, pie.height = chart_sizes.get("Overall Status", (9.0, 6.0))
    ws.add_chart(pie, "J4")

    bar = BarChart()
    chart_data = Reference(ws, min_col=2, max_col=5, min_row=module_start + 1, max_row=module_start + 1 + len(module_counts))
    cats = Reference(ws, min_col=1, min_row=module_start + 2, max_row=module_start + 1 + len(module_counts))
    bar.add_data(chart_data, titles_from_data=True)
    bar.set_categories(cats)
    bar.type = "bar"
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.title = "Test Contribution by Module"
    bar.y_axis.title = "Module"
    bar.x_axis.title = "Testcases"
    bar.width, bar.height = chart_sizes.get("Test Contribution", (15.0, 8.5))
    ws.add_chart(bar, "J20")

    ws.freeze_panes = "A4"


def format_detail_sheet(ws, rows: List[Dict[str, str]], table_name: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

    for col_idx, header in enumerate(STANDARD_HEADERS, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(bold=True, color=PALETTE["white"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX_BORDER

    status_col = STANDARD_HEADERS.index("Status") + 1
    priority_col = STANDARD_HEADERS.index("Priority") + 1
    risk_col = STANDARD_HEADERS.index("Risk Level") + 1

    for row_idx in range(2, len(rows) + 2):
        row_fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 else "F8FBFD")
        for col_idx in range(1, len(STANDARD_HEADERS) + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = copy(row_fill)
            cell.border = BOX_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        status = ws.cell(row_idx, status_col).value
        fill, font = STATUS_COLORS.get(status, ("FFFFFF", PALETTE["black"]))
        status_cell = ws.cell(row_idx, status_col)
        status_cell.fill = PatternFill("solid", fgColor=fill)
        status_cell.font = Font(bold=True, color=font)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in (priority_col, risk_col):
            value = ws.cell(row_idx, col_idx).value
            fill, font = PRIORITY_COLORS.get(value, ("FFFFFF", PALETTE["black"]))
            cell = ws.cell(row_idx, col_idx)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(bold=True, color=font)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 22,
        "B": 22,
        "C": 13,
        "D": 42,
        "E": 36,
        "F": 48,
        "G": 36,
        "H": 52,
        "I": 12,
        "J": 13,
        "K": 46,
        "L": 18,
        "M": 15,
        "N": 18,
        "O": 15,
        "P": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_idx in range(2, min(len(rows) + 2, 250)):
        ws.row_dimensions[row_idx].height = 48

    if rows:
        table_range = f"A1:{get_column_letter(len(STANDARD_HEADERS))}{len(rows) + 1}"
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)


def add_rows_sheet(wb: Workbook, sheet_name: str, rows: List[Dict[str, str]], table_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(STANDARD_HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in STANDARD_HEADERS])
    format_detail_sheet(ws, rows, table_name)


def add_bug_reports_sheet(wb: Workbook, group_name: str, reports: List[Dict[str, str]]) -> None:
    ws = wb.create_sheet("Bug Reports", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    apply_title_style(
        ws,
        f"{group_name.upper()} BUG REPORTS",
        "Detailed defect and requirement-gap list generated from non-passed testcase rows.",
    )

    summary = Counter(report["Issue Type"] for report in reports)
    severity_summary = Counter(report["Severity"] for report in reports)
    write_kpi_card(ws, "A4:B5", "TOTAL ISSUES", len(reports), PALETTE["navy"])
    write_kpi_card(ws, "C4:D5", "BUGS", summary.get("Bug", 0), PALETTE["red"])
    write_kpi_card(ws, "E4:F5", "GAPS", summary.get("Requirement Gap", 0) + summary.get("Test Gap", 0), PALETTE["orange"])
    write_kpi_card(ws, "G4:H5", "HIGH+", severity_summary.get("High", 0) + severity_summary.get("Critical", 0), PALETTE["blue"])

    start_row = 8
    for col_idx, header in enumerate(BUG_REPORT_HEADERS, start=1):
        cell = ws.cell(start_row, col_idx, header)
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(bold=True, color=PALETTE["white"])
        cell.border = BOX_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, report in enumerate(reports, start=start_row + 1):
        for col_idx, header in enumerate(BUG_REPORT_HEADERS, start=1):
            cell = ws.cell(row_idx, col_idx, report.get(header, ""))
            cell.border = BOX_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 else "F8FBFD")

        issue_type = report.get("Issue Type", "")
        fill, font = ISSUE_TYPE_COLORS.get(issue_type, ("FFFFFF", PALETTE["black"]))
        issue_cell = ws.cell(row_idx, BUG_REPORT_HEADERS.index("Issue Type") + 1)
        issue_cell.fill = PatternFill("solid", fgColor=fill)
        issue_cell.font = Font(bold=True, color=font)
        issue_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        severity = report.get("Severity", "")
        fill, font = SEVERITY_COLORS.get(severity, ("FFFFFF", PALETTE["black"]))
        severity_cell = ws.cell(row_idx, BUG_REPORT_HEADERS.index("Severity") + 1)
        severity_cell.fill = PatternFill("solid", fgColor=fill)
        severity_cell.font = Font(bold=True, color=font)
        severity_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        status = report.get("Status", "")
        fill, font = STATUS_COLORS.get(status, ("FFFFFF", PALETTE["black"]))
        status_cell = ws.cell(row_idx, BUG_REPORT_HEADERS.index("Status") + 1)
        status_cell.fill = PatternFill("solid", fgColor=fill)
        status_cell.font = Font(bold=True, color=font)
        status_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 18,
        "B": 18,
        "C": 13,
        "D": 10,
        "E": 12,
        "F": 24,
        "G": 24,
        "H": 44,
        "I": 48,
        "J": 56,
        "K": 56,
        "L": 54,
        "M": 58,
        "N": 58,
        "O": 48,
        "P": 32,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_idx in range(start_row + 1, start_row + 1 + len(reports)):
        ws.row_dimensions[row_idx].height = 70

    if reports:
        table_range = f"A{start_row}:{get_column_letter(len(BUG_REPORT_HEADERS))}{start_row + len(reports)}"
        table = Table(displayName=f"tbl_{group_name.title()}BugReports", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    else:
        ws.cell(start_row + 1, 1, "No failed/skipped/not-run issue rows detected for this group.")
        ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=8)


def add_bug_report_guide(wb: Workbook, group_name: str) -> None:
    ws = wb.create_sheet("Bug Report Guide")
    ws.sheet_view.showGridLines = False
    apply_title_style(
        ws,
        "BUG REPORT GUIDE",
        f"Checklist for reporting {group_name.title()} defects to developers.",
    )
    sections = [
        ("Bug ID", "BUG-<module>-<sequence>, ví dụ BUG-SEARCH-001"),
        ("Title", "Viết ngắn, có điều kiện + kết quả sai. Ví dụ: Không giới hạn số ghế theo số vé khi đặt chuyến."),
        ("Environment", "URL, branch/build, browser, viewport, account/role, backend env, date/time chạy test."),
        ("Related Testcase", "Dẫn TC ID, module, file CSV/workbook, Playwright spec nếu có."),
        ("Steps To Reproduce", "Liệt kê từng bước cụ thể, không gộp nhiều hành động vào một dòng."),
        ("Actual Result", "Hệ thống đang làm gì. Ghi rõ message, URL, API response, trạng thái DB nếu có."),
        ("Expected Result", "Hệ thống nên làm gì theo requirement hoặc hành vi đã thống nhất."),
        ("Evidence", "Screenshot/video, Playwright trace, console/network log, API payload/response, DB record liên quan."),
        ("Severity", "Critical: mất tiền/mất dữ liệu; High: chặn luồng chính; Medium: sai chức năng nhưng có workaround; Low: UI/copy nhỏ."),
        ("Priority", "P0/P1/P2/P3 theo mức cần sửa sớm, tách khỏi severity nếu bug nghiêm trọng nhưng ít gặp."),
        ("Root Cause Hint", "Không bắt buộc, chỉ ghi nếu đã thấy code/API liên quan. Tránh đoán chắc khi chưa có bằng chứng."),
        ("Regression Scope", "Gợi ý testcase cần chạy lại sau khi dev sửa."),
    ]
    ws["A4"] = "Field"
    ws["B4"] = "What to write"
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(bold=True, color=PALETTE["white"])
        cell.border = BOX_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, (field, guidance) in enumerate(sections, start=5):
        ws.cell(idx, 1, field)
        ws.cell(idx, 2, guidance)
        for col in (1, 2):
            cell = ws.cell(idx, col)
            cell.border = BOX_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFD")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    for row in range(5, 5 + len(sections)):
        ws.row_dimensions[row].height = 42
    ws.freeze_panes = "A5"


def add_readme_sheet(wb: Workbook, group_name: str, source_files: List[Path]) -> None:
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    apply_title_style(ws, "WORKBOOK README", f"Source files included in {group_name.title()} consolidated testcase workbook.")
    rows = [["#", "Source File"], *[[idx, path.name] for idx, path in enumerate(source_files, start=1)]]
    for row in rows:
        ws.append(row)
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=PALETTE["navy"])
        cell.font = Font(bold=True, color=PALETTE["white"])
        cell.border = BOX_BORDER
    for row_idx in range(5, 5 + len(source_files)):
        for col in (1, 2):
            ws.cell(row_idx, col).border = BOX_BORDER
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55


def save_workbook_with_fallback(wb: Workbook, output_path: Path) -> Path:
    try:
        wb.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(
            f"{output_path.stem}_{datetime.now().strftime('%H%M%S')}{output_path.suffix}"
        )
        wb.save(fallback)
        return fallback


def md_text(value: str) -> str:
    return (value or "").replace("\r\n", "\n").replace("\r", "\n").strip()


def write_markdown_issue(lines: List[str], report: Dict[str, str]) -> None:
    lines.extend(
        [
            f"### {report['Bug ID']} - {report['Title']}",
            f"- Issue Type: {report['Issue Type']}",
            f"- Severity / Priority: {report['Severity']} / {report['Priority']}",
            f"- Status in testcase: {report['Status']}",
            f"- Module: {report['Module']}",
            f"- Related testcase: {report['Related TC']} (`{report['Source File']}`)",
            f"- Suspected area: `{report['Suspected Area']}`",
            "",
            "**Steps / Trigger**",
            md_text(report["Steps / Trigger"]) or "Not captured in testcase CSV.",
            "",
            "**Actual Result**",
            md_text(report["Actual Result"]) or "Not captured in testcase CSV.",
            "",
            "**Expected Result**",
            md_text(report["Expected Result"]) or "Not captured in testcase CSV.",
            "",
            "**Evidence**",
            md_text(report["Evidence"]),
            "",
            "**Recommendation**",
            md_text(report["Recommendation"]),
            "",
            "**Regression Scope**",
            md_text(report["Regression Scope"]),
            "",
        ]
    )


def write_detailed_bug_report(
    admin_rows: List[Dict[str, str]],
    customer_rows: List[Dict[str, str]],
    admin_reports: List[Dict[str, str]],
    customer_reports: List[Dict[str, str]],
) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = REPORTS_DIR / f"txp_bug_report_detailed_{datetime.now().strftime('%Y-%m-%d')}.md"

    all_reports = admin_reports + customer_reports
    all_rows = admin_rows + customer_rows
    status_totals = Counter(row["Status"] for row in all_rows)
    type_totals = Counter(report["Issue Type"] for report in all_reports)
    severity_totals = Counter(report["Severity"] for report in all_reports)

    lines: List[str] = [
        "# TXP Bus - Detailed Bug Report",
        "",
        f"- Generated at: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "- Design By: Đỗ Thị Phương",
        "- Execute By: Playwright",
        "- Evidence source: consolidated CSV testcase status, Playwright E2E notes, and code review of `backend/` + `kien-tap/`.",
        "",
        "## Executive Summary",
        f"- Total testcase rows reviewed: {len(all_rows)}",
        f"- Passed: {status_totals.get('Passed', 0)}",
        f"- Failed: {status_totals.get('Failed', 0)}",
        f"- Skip: {status_totals.get('Skip', 0)}",
        f"- Not Run: {status_totals.get('Not Run', 0)}",
        f"- Reported issues/gaps: {len(all_reports)}",
        f"- Bugs: {type_totals.get('Bug', 0)}",
        f"- Requirement gaps: {type_totals.get('Requirement Gap', 0)}",
        f"- Test gaps: {type_totals.get('Test Gap', 0)}",
        f"- High/Critical items: {severity_totals.get('High', 0) + severity_totals.get('Critical', 0)}",
        "",
        "## Risk Focus",
        "- Fix first: seat/booking validation, held-seat handling, dispatch resource conflicts, and blacklist input validation.",
        "- Medium priority: OTP request state, location search normalization, and UX behaviors that can cause duplicated actions or confusing flows.",
        "- Test follow-up: skipped/not-run search-trip scenarios should be revisited after product rules are confirmed.",
        "",
    ]

    for label, rows, reports in [
        ("Admin", admin_rows, admin_reports),
        ("Customer", customer_rows, customer_reports),
    ]:
        counts = Counter(row["Status"] for row in rows)
        issue_counts = Counter(report["Issue Type"] for report in reports)
        lines.extend(
            [
                f"## {label} Summary",
                f"- Testcase rows: {len(rows)}",
                f"- Passed / Failed / Skip / Not Run: {counts.get('Passed', 0)} / {counts.get('Failed', 0)} / {counts.get('Skip', 0)} / {counts.get('Not Run', 0)}",
                f"- Issues in report: {len(reports)}",
                f"- Bug / Requirement Gap / Test Gap: {issue_counts.get('Bug', 0)} / {issue_counts.get('Requirement Gap', 0)} / {issue_counts.get('Test Gap', 0)}",
                "",
            ]
        )

    lines.append("## Detailed Admin Issues")
    if admin_reports:
        for report in admin_reports:
            write_markdown_issue(lines, report)
    else:
        lines.extend(["No admin issues detected from testcase CSV status.", ""])

    lines.append("## Detailed Customer Issues")
    if customer_reports:
        for report in customer_reports:
            write_markdown_issue(lines, report)
    else:
        lines.extend(["No customer issues detected from testcase CSV status.", ""])

    ticket_lookup_rows = [row for row in customer_rows if row.get("Source File") == "testcases_customer_ticket_lookup.csv"]
    ticket_lookup_not_run = sum(1 for row in ticket_lookup_rows if row.get("Status") == "Not Run")
    if ticket_lookup_not_run:
        lines.extend(
            [
                "## Coverage Note",
                f"- `testcases_customer_ticket_lookup.csv` has {ticket_lookup_not_run} rows with blank/Not Run status. This is tracked as execution coverage debt, not as individual product bugs.",
                "- Recommendation: run or split ticket lookup cases before final QA sign-off so the dashboard no longer hides a full module as unexecuted.",
                "",
            ]
        )

    lines.extend(
        [
            "## Suggested Dev Ticket Format",
            "Use one ticket per `Bug ID` and keep the title format: `[Severity][Module] Short failing behavior`.",
            "",
            "Required fields:",
            "- Environment: branch/build, frontend URL, backend URL, browser, viewport, test account/role.",
            "- Related testcase: TC ID + CSV file + Playwright spec if available.",
            "- Steps to reproduce: copy from this report, then add exact test data if the dev needs it.",
            "- Actual result: observed UI/API/DB behavior.",
            "- Expected result: business rule or agreed behavior.",
            "- Evidence: HTML report, screenshot/video/trace, API payload/response, console/network logs.",
            "- Acceptance criteria: bug is fixed, affected Playwright case passes, and linked regression scope passes.",
            "",
        ]
    )

    output_path.write_text("\n".join(lines), encoding="utf-8")
    return output_path


def build_workbook(group_name: str, source_files: List[Path], output_name: str) -> tuple[Path, List[Dict[str, str]], List[Dict[str, str]]]:
    rows = load_group_rows(group_name.lower(), source_files)
    bug_reports = build_bug_reports(group_name, rows)
    chart_sizes = copy_template_chart_sizes(TEMPLATE_PATH)
    wb = Workbook()
    wb.properties.creator = "Đỗ Thị Phương"
    wb.properties.title = f"TXP Bus {group_name.title()} Testcases"
    add_dashboard(wb, group_name, rows, chart_sizes)
    add_bug_reports_sheet(wb, group_name, bug_reports)

    used = {"Dashboard"}
    add_rows_sheet(wb, f"All {group_name.title()}", rows, f"tbl_All_{group_name.title().replace(' ', '')}")
    used.add(f"All {group_name.title()}")

    grouped: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row["Source File"]].append(row)

    for path in source_files:
        module_rows = grouped[path.name]
        if not module_rows:
            continue
        base_name = pretty_module_name(path, group_name.lower())
        sheet_name = safe_sheet_name(base_name, used)
        table_base = "".join(ch for ch in sheet_name if ch.isalnum())[:20] or "Module"
        add_rows_sheet(wb, sheet_name, module_rows, f"tbl_{group_name.title()}{table_base}")

    add_bug_report_guide(wb, group_name)
    add_readme_sheet(wb, group_name, source_files)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / output_name
    output_path = save_workbook_with_fallback(wb, output_path)
    return output_path, bug_reports, rows


def validate_workbook(path: Path) -> Dict[str, int]:
    wb = load_workbook(path)
    chart_count = sum(len(ws._charts) for ws in wb.worksheets)
    table_count = sum(len(ws.tables) for ws in wb.worksheets)
    return {
        "sheets": len(wb.worksheets),
        "charts": chart_count,
        "tables": table_count,
    }


def main() -> None:
    admin_files = sorted(TESTCASE_DIR.glob("testcases_admin_*.csv"))
    customer_files = sorted(TESTCASE_DIR.glob("testcases_customer_*.csv"))
    search_trip = TESTCASE_DIR / "testcases_search_trip.csv"
    if search_trip.exists():
        customer_files.append(search_trip)

    date_suffix = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    admin_output, admin_reports, admin_rows = build_workbook(
        "Admin",
        admin_files,
        f"txp_admin_testcases_consolidated_{date_suffix}.xlsx",
    )
    customer_output, customer_reports, customer_rows = build_workbook(
        "Customer",
        customer_files,
        f"txp_customer_testcases_consolidated_{date_suffix}.xlsx",
    )
    bug_report = write_detailed_bug_report(admin_rows, customer_rows, admin_reports, customer_reports)

    outputs = [admin_output, customer_output]
    for output in outputs:
        info = validate_workbook(output)
        print(f"{output} | sheets={info['sheets']} charts={info['charts']} tables={info['tables']}")
    print(f"{bug_report} | issues={len(admin_reports) + len(customer_reports)}")


if __name__ == "__main__":
    main()
